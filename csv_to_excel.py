import sys
import io
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

try:
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
except ImportError as e:
    print(f"[LOI] Thieu thu vien: {e}")
    print("Chay: pip install pandas openpyxl")
    sys.exit(1)

CSV_DIR = Path("csv")
OUTPUT = Path("tnbike_export.xlsx")

TABLES = [
    "product_group",
    "product_line",
    "product",
    "product_price",
    "province",
    "customer",
    "sales_order",
    "order_line",
    "fact_sales",
]

HEADER_FILL  = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT  = Font(color="FFFFFF", bold=True)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")


def find_csv(table: str) -> Path | None:
    matches = sorted(CSV_DIR.glob(f"{table}*.csv"))
    return matches[0] if matches else None


def style_sheet(ws):
    for cell in ws[1]:
        cell.fill  = HEADER_FILL
        cell.font  = HEADER_FONT
        cell.alignment = HEADER_ALIGN

    for col_idx, col_cells in enumerate(ws.columns, 1):
        width = max(
            (len(str(c.value)) if c.value is not None else 0) for c in col_cells
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(width + 4, 60)

    ws.freeze_panes = "A2"


def main():
    if not CSV_DIR.exists():
        print(f"[LOI] Thu muc '{CSV_DIR}' khong ton tai.")
        sys.exit(1)

    print(f"Dang doc CSV tu: {CSV_DIR}/")
    print(f"{'Bang':<20} {'So dong':>10}")
    print("-" * 32)

    dfs: dict[str, pd.DataFrame] = {}
    for table in TABLES:
        csv_path = find_csv(table)
        if csv_path is None:
            print(f"  {table:<20} {'(khong tim thay)':>16}")
            continue
        df = pd.read_csv(csv_path, encoding="utf-8-sig")
        dfs[table] = df
        print(f"  {table:<20} {len(df):>10,} dong")

    if not dfs:
        print("[LOI] Khong co file CSV nao hop le.")
        sys.exit(1)

    print("-" * 32)
    print(f"  {'TONG':<20} {sum(len(d) for d in dfs.values()):>10,} dong\n")

    print(f"Dang ghi Excel: {OUTPUT}")
    with pd.ExcelWriter(OUTPUT, engine="openpyxl") as writer:
        for table, df in dfs.items():
            df.to_excel(writer, sheet_name=table, index=False)

    wb = load_workbook(OUTPUT)
    for sheet in wb.sheetnames:
        style_sheet(wb[sheet])
    wb.save(OUTPUT)

    print(f"Hoan tat! => {OUTPUT}  ({len(dfs)} sheets)")


if __name__ == "__main__":
    main()
