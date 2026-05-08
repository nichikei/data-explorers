import sys
import io
from datetime import datetime

# Fix Unicode output on Windows terminal
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")
from pathlib import Path

try:
    import pandas as pd
    from sqlalchemy import create_engine, text
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
except ImportError as e:
    print(f"[LOI] Thieu thu vien: {e}")
    print("Chay: pip install psycopg2-binary pandas openpyxl sqlalchemy")
    sys.exit(1)

# ── Cấu hình kết nối ──────────────────────────────────────────────────────────
DB_CONFIG = {
    "host": "localhost",
    "port": 5432,
    "dbname": "tnbike_db",
    "user": "postgres",
    "password": "1",
}
SCHEMA = "tnbike"
TABLES = [
    "product_group",
    "product_line",
    "product",
    "product_price",
    "province",
    "customer",
    "sales_order",
    "order_line",
    "fact_sales",
    "email_log",
]

# ── Đường dẫn output ──────────────────────────────────────────────────────────
TODAY = datetime.now().strftime("%Y%m%d")
EXCEL_FILE = Path(f"tnbike_export_{TODAY}.xlsx")
CSV_DIR = Path("csv")
CSV_DIR.mkdir(exist_ok=True)

# ── Màu header Excel ──────────────────────────────────────────────────────────
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")


def connect():
    try:
        url = (
            f"postgresql+psycopg2://{DB_CONFIG['user']}:{DB_CONFIG['password']}"
            f"@{DB_CONFIG['host']}:{DB_CONFIG['port']}/{DB_CONFIG['dbname']}"
        )
        engine = create_engine(url)
        with engine.connect() as c:
            c.execute(text("SELECT 1"))
        print(f"[OK] Ket noi thanh cong: {DB_CONFIG['host']}:{DB_CONFIG['port']}/{DB_CONFIG['dbname']}")
        return engine
    except Exception as e:
        print(f"[LOI] Khong the ket noi database:\n  {e}")
        print("Kiem tra lai host, port, user, password va ten database.")
        sys.exit(1)


def strip_tz(df: pd.DataFrame) -> pd.DataFrame:
    for col in df.select_dtypes(include=["datetimetz"]).columns:
        df[col] = df[col].dt.tz_localize(None)
    return df


def fetch_table(engine, table: str) -> pd.DataFrame:
    query = f'SELECT * FROM {SCHEMA}."{table}"'
    try:
        df = pd.read_sql(query, engine)
        return strip_tz(df)
    except Exception as e:
        print(f"  [WARN] Bo qua bang '{table}': {e}")
        return pd.DataFrame()


def style_sheet(ws):
    """Áp dụng màu header, tự động căn độ rộng cột, freeze row đầu."""
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN

    for col_idx, col_cells in enumerate(ws.columns, 1):
        max_len = max(
            (len(str(c.value)) if c.value is not None else 0) for c in col_cells
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 60)

    ws.freeze_panes = "A2"


def main():
    engine = connect()
    print()

    dfs: dict[str, pd.DataFrame] = {}
    for table in TABLES:
        df = fetch_table(engine, table)
        dfs[table] = df

    engine.dispose()

    # ── Ghi Excel ─────────────────────────────────────────────────────────────
    print(f"Đang ghi Excel: {EXCEL_FILE}")
    with pd.ExcelWriter(EXCEL_FILE, engine="openpyxl") as writer:
        for table, df in dfs.items():
            if df.empty:
                continue
            df.to_excel(writer, sheet_name=table, index=False)

    wb = load_workbook(EXCEL_FILE)
    for sheet_name in wb.sheetnames:
        style_sheet(wb[sheet_name])
    wb.save(EXCEL_FILE)
    print(f"  => Đã lưu: {EXCEL_FILE}\n")

    # ── Ghi CSV + in thống kê ─────────────────────────────────────────────────
    print(f"Đang ghi CSV vào thư mục: {CSV_DIR}/")
    print(f"\n{'Bảng':<20} {'Số dòng':>10}")
    print("-" * 32)
    for table, df in dfs.items():
        if df.empty:
            print(f"  {table:<20} {'(bỏ qua)':>10}")
            continue
        csv_path = CSV_DIR / f"{table}_{TODAY}.csv"
        df.to_csv(csv_path, index=False, encoding="utf-8-sig")
        print(f"  {table:<20} {len(df):>10,} dòng")

    print("-" * 32)
    total = sum(len(df) for df in dfs.values())
    print(f"  {'TỔNG':<20} {total:>10,} dòng")
    print(f"\nHoàn tất! Excel: {EXCEL_FILE} | CSV: {CSV_DIR}/")


if __name__ == "__main__":
    main()
